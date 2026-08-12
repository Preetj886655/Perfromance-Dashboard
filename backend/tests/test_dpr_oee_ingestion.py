"""Integration tests for DPR_OEE Excel ingestion (service layer).

Uses Compose Postgres (127.0.0.1:5433 / pril_analytics) inside a rolled-back
transaction so no temporary masters/production/import rows remain.

Covers user validation items 1–20 (21–23 are regression suites run separately).
"""

from __future__ import annotations

import io
import uuid
from datetime import date, datetime, time
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.db.session import get_engine
from app.models.downtime_event import DowntimeEvent
from app.models.downtime_reason import DowntimeReason
from app.models.import_job import ImportJob
from app.models.import_job_row import ImportJobRow
from app.models.machine import Machine
from app.models.machine_status import MachineStatus
from app.models.machine_type import MachineType
from app.models.operator import Operator
from app.models.part import Part
from app.models.plant import Plant
from app.models.production_record import ProductionRecord
from app.models.production_record_metrics import ProductionRecordMetrics
from app.models.rejection_event import RejectionEvent
from app.models.rejection_reason import RejectionReason
from app.models.shift import Shift
from app.services.dpr_oee_ingestion import (
    DOWNTIME_COLUMNS,
    REJECTION_COLUMNS,
    SHEET_NAME,
    ingest_dpr_oee_workbook,
    validate_dpr_oee_sheet,
)
from app.services.oee_calculator import calculate_oee_metrics

REPO_ROOT = Path(__file__).resolve().parents[2]
REAL_XLSX = REPO_ROOT / "PRIL_DPR_OEE Sheet - PG_NPD_029.xlsx"

DOWNTIME_SEED: tuple[tuple[str, str, int, str], ...] = (
    ("1", "Manpower Shortage", 1, "Q"),
    ("2", "Mould Trial", 2, "R"),
    ("3", "Bin Shortage", 3, "S"),
    ("4", "Material Shortage", 4, "T"),
    ("5", "M/c Under BD", 5, "U"),
    ("6", "Nozzle Block", 6, "V"),
    ("7", "Mould Problem", 7, "W"),
    ("8", "Crystal/ Insert Shortage", 8, "X"),
    ("9", "Power Failure", 9, "Y"),
    ("10", "Process Setting", 10, "Z"),
    ("11", "Others", 11, "AA"),
)


@pytest.fixture
def db_session() -> Session:
    """Session bound to an outer transaction that always rolls back."""
    engine = get_engine()
    connection = engine.connect()
    transaction = connection.begin()
    session = Session(bind=connection, autoflush=False, expire_on_commit=False)
    try:
        yield session
    finally:
        session.close()
        transaction.rollback()
        connection.close()


def _uid(prefix: str) -> str:
    return f"{prefix}_{uuid.uuid4().hex[:12]}"


def _seed_downtime_reasons(session: Session) -> dict[str, DowntimeReason]:
    """Transactional fixture seed only — not a permanent production seed."""
    by_col: dict[str, DowntimeReason] = {}
    for code, label, sort_order, excel_column in DOWNTIME_SEED:
        existing = session.scalar(
            select(DowntimeReason).where(DowntimeReason.code == code)
        )
        if existing is None:
            row = DowntimeReason(
                code=code,
                label=label,
                category="unplanned",  # illustrative only; Q2 TBC
                is_active=True,
                sort_order=sort_order,
                excel_column=excel_column,
            )
            session.add(row)
            session.flush()
            by_col[excel_column] = row
        else:
            if not existing.excel_column:
                existing.excel_column = excel_column
            by_col[excel_column] = existing
    session.flush()
    return by_col


def _ensure_rejection_reasons(session: Session) -> None:
    """Assert A–J exist (already seeded in DB); skip inventing if present."""
    codes = set(
        session.scalars(select(RejectionReason.code)).all()
    )
    missing = {c for _, c in REJECTION_COLUMNS} - codes
    assert not missing, (
        f"rejection_reasons A–J must be seeded before ingestion tests; missing={missing}"
    )


def _seed_masters_for_real_xlsx(session: Session) -> dict[str, object]:
    """Plant + masters matching Excel rows 5–6 (M001 / A / PD001 / ABC)."""
    _ensure_rejection_reasons(session)
    _seed_downtime_reasons(session)

    plant = Plant(
        code=_uid("PLT"),
        name="DPR Ingest Test Plant",
        timezone="Asia/Kolkata",
        is_active=True,
    )
    session.add(plant)
    session.flush()

    mtype = MachineType(code=_uid("MT"), name="Injection", is_active=True)
    mstatus = MachineStatus(code=_uid("MS"), name="Active", is_active=True)
    session.add_all([mtype, mstatus])
    session.flush()

    machine = Machine(
        plant_id=plant.id,
        code="M001",
        name="Machine 001",
        machine_type_id=mtype.id,
        status_id=mstatus.id,
    )
    shift = Shift(
        plant_id=plant.id,
        code="A",
        name="Shift A",
        start_time=time(8, 30),
        end_time=time(20, 30),
        crosses_midnight=False,
    )
    part = session.scalar(select(Part).where(Part.code == "PD001"))
    if part is None:
        part = Part(code="PD001", name="RGP")
        session.add(part)
    operator = session.scalar(select(Operator).where(Operator.name == "ABC"))
    if operator is None:
        operator = Operator(employee_code=_uid("EMP"), name="ABC")
        session.add(operator)

    session.add_all([machine, shift])
    session.flush()

    return {
        "plant": plant,
        "machine": machine,
        "shift": shift,
        "part": part,
        "operator": operator,
    }


def _write_minimal_workbook(path_or_buf, *, rows: list[dict]) -> None:
    """Build a DPR_OEE-shaped workbook for synthetic cases."""
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    # Title / headers matching real sheet closely enough for validator.
    ws["A1"] = "PATIL RAIL INFRASTRUCTURE PVT. LTD."
    ws["A2"] = "DAILY PRODUCTION REPORT WITH OEE"
    headers = {
        "A": "S.No.",
        "B": "Date",
        "C": "Shift",
        "D": "Machine Name/No.",
        "E": "Start Time",
        "F": "Stop Time",
        "G": "Shift Time (Minutes)",
        "H": "Operator Name",
        "I": "Part Name ",
        "J": "Part No.",
        "K": "Cavity",
        "L": "Cycle Time (Sec.)",
        "M": "Target Qty./Hr. (Pcs.)",
        "N": "Prod. Qty. (Pcs.)",
        "O": "Planned Down Time (Tea/Lunch)",
        "P": "Available Time",
        "Q": "Reason of Idle Time (Unplanned BD Time in Minutes)",
        "AB": "Total Idle Time (Minutes)",
        "AC": "Total Run Time (Minutes)",
        "AD": "Availability Ratio (A)",
        "AE": "Actual Qty./ Hr.",
        "AF": "Operator Efficiency (Performance Ratio) - (P)",
        "AG": "Machine Efficiency (Machine Utilisation)",
        "AH": "Reason of  Rejection (Qty. in Pcs.)",
        "AR": "Total Rejection (Pcs Qty.)",
        "AS": "Rejection PPM",
        "AT": "Quantity Ratio (Q)",
        "AU": "OEE (A*P*Q)",
        "AV": "Any Other Remarks",
    }
    for col, val in headers.items():
        ws[f"{col}3"] = val
    for letter, code, label in DOWNTIME_COLUMNS:
        ws[f"{letter}4"] = f"{code}. {label}"
    for letter, code in REJECTION_COLUMNS:
        ws[f"{letter}4"] = f"{code}. Reason"

    for i, row in enumerate(rows):
        r = 5 + i
        for col, val in row.items():
            ws[f"{col}{r}"] = val
        # Formula placeholders like real template (ignored by ingestion).
        ws[f"G{r}"] = f'=IFERROR((F{r}-E{r})*24*60,"")'
        ws[f"M{r}"] = f'=IFERROR(3600/(L{r}/K{r}),"")'

    # Extra empty template row with formulas only (like rows 7–30).
    empty_r = 5 + len(rows)
    ws[f"G{empty_r}"] = f'=IFERROR((F{empty_r}-E{empty_r})*24*60,"")'
    ws[f"M{empty_r}"] = f'=IFERROR(3600/(L{empty_r}/K{empty_r}),"")'

    if isinstance(path_or_buf, str | Path):
        wb.save(path_or_buf)
    else:
        wb.save(path_or_buf)


# --- 1–3: sheet discovery, headers, empty skip ---


def test_sheet_discovery_and_header_mapping(db_session: Session) -> None:
    assert REAL_XLSX.exists()
    wb = load_workbook(REAL_XLSX, data_only=False)
    assert SHEET_NAME in wb.sheetnames
    errors = validate_dpr_oee_sheet(wb[SHEET_NAME])
    assert errors == []


def test_wrong_sheet_fails_job(db_session: Session, tmp_path: Path) -> None:
    masters = _seed_masters_for_real_xlsx(db_session)
    plant: Plant = masters["plant"]  # type: ignore[assignment]
    path = tmp_path / "bad.xlsx"
    wb = Workbook()
    wb.active.title = "NOT_DPR"
    wb.save(path)
    result = ingest_dpr_oee_workbook(db_session, path, plant_id=plant.id)
    assert result.status == "failed"
    assert "DPR_OEE" in (result.error_summary or "")


def test_empty_template_rows_skipped(db_session: Session) -> None:
    masters = _seed_masters_for_real_xlsx(db_session)
    plant: Plant = masters["plant"]  # type: ignore[assignment]
    result = ingest_dpr_oee_workbook(db_session, REAL_XLSX, plant_id=plant.id)
    assert result.status == "committed"
    assert result.success_count == 2  # rows 5–6 processed (shared business key → upsert)
    assert result.skipped_count >= 20  # formulas-only template rows
    job_rows = db_session.scalars(
        select(ImportJobRow).where(ImportJobRow.import_job_id == result.import_job_id)
    ).all()
    assert len(job_rows) == 2
    assert {r.row_number for r in job_rows} == {5, 6}


def _seed_second_machine(session: Session, masters: dict[str, object]) -> Machine:
    plant: Plant = masters["plant"]  # type: ignore[assignment]
    machine: Machine = masters["machine"]  # type: ignore[assignment]
    m2 = Machine(
        plant_id=plant.id,
        code="M002",
        name="Machine 002",
        machine_type_id=machine.machine_type_id,
        status_id=machine.status_id,
    )
    session.add(m2)
    session.flush()
    return m2


def _row5_cells(*, machine: str = "M001") -> dict:
    return {
        "B": datetime(2026, 8, 8),
        "C": "A",
        "D": machine,
        "E": time(8, 30),
        "F": time(20, 30),
        "H": "ABC",
        "I": "RGP",
        "J": "PD001",
        "K": 2,
        "L": 60,
        "N": 1200,
        "O": 60,
        "U": 20,  # M/c Under BD
        "AH": 1,
        "AI": 2,
        "AJ": 3,
        "AK": 5,
        "AL": 4,
    }


def _row6_cells(*, machine: str = "M001") -> dict:
    return {
        "B": datetime(2026, 8, 8),
        "C": "A",
        "D": machine,
        "E": time(8, 30),
        "F": time(20, 30),
        "H": "ABC",
        "I": "RGP",
        "J": "PD001",
        "K": 2,
        "L": 60,
        "N": 1100,
        "O": 30,
        "S": 20,  # Bin Shortage
        "AM": 4,  # Dent Mark
    }


# --- 4–16: happy path row 5/6 ---


def test_ingest_real_xlsx_duplicate_business_key_last_wins(
    db_session: Session,
) -> None:
    """Real sample rows 5–6 share machine+shift+date+part+start — schema upserts.

    Unique (machine, shift, production_date, part, start_at) + external_row_key
    mean both Excel sample rows map to one production_record (last write wins).
    Not a schema gap for real ops (duplicate keys are invalid shop-floor data);
    OEE parity for both fixtures is covered with distinct-key synthetic sheets.
    """
    masters = _seed_masters_for_real_xlsx(db_session)
    plant: Plant = masters["plant"]  # type: ignore[assignment]

    result = ingest_dpr_oee_workbook(db_session, REAL_XLSX, plant_id=plant.id)
    assert result.status == "committed"
    assert result.success_count == 2
    assert result.error_count == 0

    job = db_session.get(ImportJob, result.import_job_id)
    assert job is not None
    assert job.source_type == "excel"

    distinct_ids = set(result.production_record_ids)
    assert len(distinct_ids) == 1
    rec = db_session.get(ProductionRecord, next(iter(distinct_ids)))
    assert rec is not None
    assert rec.source_type == "excel"
    assert rec.source_import_id == result.import_job_id
    assert rec.plant_id == plant.id
    assert rec.production_date == date(2026, 8, 8)
    # Last processed sample row (row 6) wins
    assert rec.produced_qty == Decimal("1100")
    assert rec.planned_downtime_min == Decimal("30")

    job_rows = list(
        db_session.scalars(
            select(ImportJobRow).where(
                ImportJobRow.import_job_id == result.import_job_id
            )
        ).all()
    )
    assert len(job_rows) == 2
    assert all(jr.production_record_id == rec.id for jr in job_rows)


def test_ingest_row5_and_row6_distinct_keys(db_session: Session, tmp_path: Path) -> None:
    """Items 4–16 with distinct business keys so both records persist."""
    masters = _seed_masters_for_real_xlsx(db_session)
    plant: Plant = masters["plant"]  # type: ignore[assignment]
    _seed_second_machine(db_session, masters)

    path = tmp_path / "rows56.xlsx"
    _write_minimal_workbook(
        path,
        rows=[_row5_cells(machine="M001"), _row6_cells(machine="M002")],
    )
    result = ingest_dpr_oee_workbook(db_session, path, plant_id=plant.id)
    assert result.status == "committed"
    assert result.success_count == 2
    assert result.error_count == 0
    assert len(set(result.production_record_ids)) == 2

    records = list(
        db_session.scalars(
            select(ProductionRecord).where(
                ProductionRecord.id.in_(result.production_record_ids)
            )
        ).all()
    )
    assert len(records) == 2
    for rec in records:
        assert rec.source_type == "excel"
        assert rec.source_import_id == result.import_job_id
        assert rec.external_row_key and rec.external_row_key.startswith("dpr_oee:")

    by_qty = {r.produced_qty: r for r in records}
    r5 = by_qty[Decimal("1200")]
    r6 = by_qty[Decimal("1100")]
    assert r5.planned_downtime_min == Decimal("60")
    assert r6.planned_downtime_min == Decimal("30")
    assert r5.cavity_count == Decimal("2")
    assert r5.cycle_time_sec == Decimal("60")

    dt5 = list(
        db_session.scalars(
            select(DowntimeEvent).where(DowntimeEvent.production_record_id == r5.id)
        ).all()
    )
    dt6 = list(
        db_session.scalars(
            select(DowntimeEvent).where(DowntimeEvent.production_record_id == r6.id)
        ).all()
    )
    assert len(dt5) == 1 and dt5[0].minutes == Decimal("20")
    assert len(dt6) == 1 and dt6[0].minutes == Decimal("20")
    reason5 = db_session.get(DowntimeReason, dt5[0].downtime_reason_id)
    reason6 = db_session.get(DowntimeReason, dt6[0].downtime_reason_id)
    assert reason5 is not None and reason5.excel_column == "U" and reason5.code == "5"
    assert reason6 is not None and reason6.excel_column == "S" and reason6.code == "3"

    rj5 = list(
        db_session.scalars(
            select(RejectionEvent).where(RejectionEvent.production_record_id == r5.id)
        ).all()
    )
    rj6 = list(
        db_session.scalars(
            select(RejectionEvent).where(RejectionEvent.production_record_id == r6.id)
        ).all()
    )
    assert len(rj5) == 5
    assert sum((e.qty for e in rj5), Decimal("0")) == Decimal("15")
    assert len(rj6) == 1 and rj6[0].qty == Decimal("4")
    codes5 = {
        db_session.get(RejectionReason, e.rejection_reason_id).code  # type: ignore[union-attr]
        for e in rj5
    }
    assert codes5 == {"A", "B", "C", "D", "E"}
    code6 = db_session.get(RejectionReason, rj6[0].rejection_reason_id)
    assert code6 is not None and code6.code == "F"

    m5 = db_session.get(ProductionRecordMetrics, r5.id)
    m6 = db_session.get(ProductionRecordMetrics, r6.id)
    assert m5 is not None and m6 is not None
    assert m5.oee == pytest.approx(Decimal("0.8977272727"), abs=Decimal("1e-8"))
    assert m6.oee == pytest.approx(Decimal("0.7942028985"), abs=Decimal("1e-8"))
    assert m5.performance != m5.machine_utilisation
    assert m5.oee != pytest.approx(
        m5.availability * m5.machine_utilisation * m5.quality, abs=Decimal("1e-8")
    )
    assert m5.oee == pytest.approx(
        m5.availability * m5.performance * m5.quality, abs=Decimal("1e-8")
    )

    job_rows = list(
        db_session.scalars(
            select(ImportJobRow).where(
                ImportJobRow.import_job_id == result.import_job_id
            )
        ).all()
    )
    assert len(job_rows) == 2
    assert all(jr.production_record_id is not None for jr in job_rows)
    assert all(jr.validation_errors == [] for jr in job_rows)


def test_rejection_aj_and_downtime_qaa_mapping(db_session: Session) -> None:
    """Item 9–10: reason catalogs resolve by excel_column."""
    masters = _seed_masters_for_real_xlsx(db_session)
    plant: Plant = masters["plant"]  # type: ignore[assignment]
    result = ingest_dpr_oee_workbook(db_session, REAL_XLSX, plant_id=plant.id)
    assert result.success_count == 2
    # All downtime reasons Q–AA exist in fixture
    cols = {r.excel_column for r in db_session.scalars(select(DowntimeReason)).all()}
    assert {c for c, _, _ in DOWNTIME_COLUMNS}.issubset(cols)
    # All rejection A–J present
    codes = set(db_session.scalars(select(RejectionReason.code)).all())
    assert {c for _, c in REJECTION_COLUMNS}.issubset(codes)


# --- 17: idempotent re-import ---


def test_idempotent_reimport_updates_not_duplicates(
    db_session: Session, tmp_path: Path
) -> None:
    masters = _seed_masters_for_real_xlsx(db_session)
    plant: Plant = masters["plant"]  # type: ignore[assignment]
    _seed_second_machine(db_session, masters)
    path = tmp_path / "idem.xlsx"
    _write_minimal_workbook(
        path,
        rows=[_row5_cells(machine="M001"), _row6_cells(machine="M002")],
    )

    first = ingest_dpr_oee_workbook(db_session, path, plant_id=plant.id)
    assert first.success_count == 2
    ids_first = set(first.production_record_ids)
    assert len(ids_first) == 2
    count_after_first = db_session.scalar(
        select(func.count()).select_from(ProductionRecord).where(
            ProductionRecord.plant_id == plant.id
        )
    )

    second = ingest_dpr_oee_workbook(db_session, path, plant_id=plant.id)
    assert second.success_count == 2
    ids_second = set(second.production_record_ids)
    assert ids_first == ids_second

    count_after_second = db_session.scalar(
        select(func.count()).select_from(ProductionRecord).where(
            ProductionRecord.plant_id == plant.id
        )
    )
    assert count_after_first == count_after_second == 2

    for pid in ids_second:
        dt_n = db_session.scalar(
            select(func.count()).select_from(DowntimeEvent).where(
                DowntimeEvent.production_record_id == pid
            )
        )
        assert dt_n == 1


# --- 18: invalid row ---


def test_invalid_unknown_machine_validation_error(
    db_session: Session, tmp_path: Path
) -> None:
    masters = _seed_masters_for_real_xlsx(db_session)
    plant: Plant = masters["plant"]  # type: ignore[assignment]
    path = tmp_path / "bad_machine.xlsx"
    _write_minimal_workbook(
        path,
        rows=[
            {
                "B": datetime(2026, 8, 8),
                "C": "A",
                "D": "UNKNOWN_MACHINE",
                "E": time(8, 30),
                "F": time(20, 30),
                "H": "ABC",
                "I": "RGP",
                "J": "PD001",
                "K": 2,
                "L": 60,
                "N": 100,
                "O": 0,
            }
        ],
    )
    result = ingest_dpr_oee_workbook(db_session, path, plant_id=plant.id)
    assert result.status == "failed"
    assert result.error_count == 1
    assert result.success_count == 0
    jr = db_session.scalar(
        select(ImportJobRow).where(ImportJobRow.import_job_id == result.import_job_id)
    )
    assert jr is not None
    assert jr.production_record_id is None
    assert any("Unknown machine" in e["message"] for e in jr.validation_errors)
    assert (
        db_session.scalar(
            select(func.count()).select_from(ProductionRecord).where(
                ProductionRecord.plant_id == plant.id
            )
        )
        == 0
    )


def test_missing_downtime_reason_is_validation_error(
    db_session: Session, tmp_path: Path
) -> None:
    """Do not invent downtime reasons when catalog empty."""
    _ensure_rejection_reasons(db_session)
    # Intentionally do NOT seed downtime reasons
    plant = Plant(
        code=_uid("PLT"),
        name="No DT Plant",
        timezone="Asia/Kolkata",
        is_active=True,
    )
    session = db_session
    session.add(plant)
    session.flush()
    mtype = MachineType(code=_uid("MT"), name="T", is_active=True)
    mstatus = MachineStatus(code=_uid("MS"), name="A", is_active=True)
    session.add_all([mtype, mstatus])
    session.flush()
    session.add_all(
        [
            Machine(
                plant_id=plant.id,
                code="M001",
                name="M1",
                machine_type_id=mtype.id,
                status_id=mstatus.id,
            ),
            Shift(
                plant_id=plant.id,
                code="A",
                name="A",
                start_time=time(8, 0),
                end_time=time(16, 0),
                crosses_midnight=False,
            ),
            Part(code=_uid("PT"), name="P"),
            Operator(employee_code=_uid("E"), name="ABC"),
        ]
    )
    session.flush()
    part = session.scalars(select(Part).order_by(Part.created_at.desc())).first()
    assert part is not None

    buf = io.BytesIO()
    _write_minimal_workbook(
        buf,
        rows=[
            {
                "B": datetime(2026, 8, 8),
                "C": "A",
                "D": "M001",
                "E": time(8, 0),
                "F": time(16, 0),
                "H": "ABC",
                "I": "P",
                "J": part.code,
                "K": 1,
                "L": 30,
                "N": 10,
                "O": 0,
                "U": 5,
            }
        ],
    )
    buf.seek(0)
    result = ingest_dpr_oee_workbook(session, buf, plant_id=plant.id)
    assert result.success_count == 0
    assert result.error_count == 1
    jr = session.scalar(
        select(ImportJobRow).where(ImportJobRow.import_job_id == result.import_job_id)
    )
    assert jr is not None
    assert any("Downtime reason" in e["message"] for e in jr.validation_errors)


# --- 19: Q1 ---


def test_q1_stop_before_start_no_plus_24h(
    db_session: Session, tmp_path: Path
) -> None:
    masters = _seed_masters_for_real_xlsx(db_session)
    plant: Plant = masters["plant"]  # type: ignore[assignment]
    path = tmp_path / "q1.xlsx"
    _write_minimal_workbook(
        path,
        rows=[
            {
                "B": datetime(2026, 8, 8),
                "C": "A",
                "D": "M001",
                "E": time(22, 0),
                "F": time(6, 0),  # stop < start same calendar date
                "H": "ABC",
                "I": "RGP",
                "J": "PD001",
                "K": 2,
                "L": 60,
                "N": 100,
                "O": 0,
            }
        ],
    )
    result = ingest_dpr_oee_workbook(db_session, path, plant_id=plant.id)
    assert result.success_count == 1
    rec = db_session.get(ProductionRecord, result.production_record_ids[0])
    assert rec is not None
    assert rec.stop_at < rec.start_at  # no +24h invent
    # Calculator agreement
    calc = calculate_oee_metrics(
        start_at=rec.start_at,
        stop_at=rec.stop_at,
        cavity_count=rec.cavity_count,
        cycle_time_sec=rec.cycle_time_sec,
        produced_qty=rec.produced_qty,
        planned_downtime_min=rec.planned_downtime_min,
    )
    assert calc.q1_midnight_unresolved is True
    assert calc.shift_time_min is None
    metrics = db_session.get(ProductionRecordMetrics, rec.id)
    assert metrics is not None
    assert metrics.shift_time_min is None
    assert metrics.oee is None


# --- 20: NULL metrics ---


def test_undefined_metrics_remain_null(
    db_session: Session, tmp_path: Path
) -> None:
    """Zero produced → rejection_ppm / quality None → SQL NULL (015)."""
    masters = _seed_masters_for_real_xlsx(db_session)
    plant: Plant = masters["plant"]  # type: ignore[assignment]
    path = tmp_path / "null_metrics.xlsx"
    _write_minimal_workbook(
        path,
        rows=[
            {
                "B": datetime(2026, 8, 8),
                "C": "A",
                "D": "M001",
                "E": time(8, 30),
                "F": time(20, 30),
                "H": "ABC",
                "I": "RGP",
                "J": "PD001",
                "K": 2,
                "L": 60,
                "N": 0,  # produced 0 → PPM/quality undefined
                "O": 60,
            }
        ],
    )
    result = ingest_dpr_oee_workbook(db_session, path, plant_id=plant.id)
    assert result.success_count == 1
    metrics = db_session.get(
        ProductionRecordMetrics, result.production_record_ids[0]
    )
    assert metrics is not None
    assert metrics.rejection_ppm is None
    assert metrics.quality is None
    assert metrics.oee is None
    assert metrics.total_rejection_qty == Decimal("0")


def test_plant_id_required_parameter(db_session: Session) -> None:
    """Q11: plant comes from caller — unknown UUID raises."""
    with pytest.raises(ValueError, match="plant_id"):
        ingest_dpr_oee_workbook(
            db_session, REAL_XLSX, plant_id=uuid.uuid4()
        )


def test_bytes_ingestion(db_session: Session) -> None:
    masters = _seed_masters_for_real_xlsx(db_session)
    plant: Plant = masters["plant"]  # type: ignore[assignment]
    data = REAL_XLSX.read_bytes()
    result = ingest_dpr_oee_workbook(db_session, data, plant_id=plant.id)
    assert result.success_count == 2


def test_no_leftover_when_rolled_back(db_session: Session, tmp_path: Path) -> None:
    """Sanity: counts visible inside txn; outer fixture rolls back."""
    masters = _seed_masters_for_real_xlsx(db_session)
    plant: Plant = masters["plant"]  # type: ignore[assignment]
    _seed_second_machine(db_session, masters)
    path = tmp_path / "rollback.xlsx"
    _write_minimal_workbook(
        path,
        rows=[_row5_cells(machine="M001"), _row6_cells(machine="M002")],
    )
    before = db_session.scalar(select(func.count()).select_from(ProductionRecord))
    ingest_dpr_oee_workbook(db_session, path, plant_id=plant.id)
    after = db_session.scalar(select(func.count()).select_from(ProductionRecord))
    assert after == before + 2
