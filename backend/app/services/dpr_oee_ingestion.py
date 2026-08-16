"""DPR_OEE Excel ingestion (service layer only).

Pipeline
--------
Excel workbook → ``import_jobs`` → sheet/header validation → map columns B–AV →
skip empty template rows → validate / resolve masters → ``production_records``
(RAW) → non-zero ``downtime_events`` / ``rejection_events`` →
``calculate_oee_metrics`` via ``persist_production_record_metrics`` →
``import_job_rows`` lineage (payload + validation_errors).

Timestamps
----------
Excel Date (B) + Start/Stop (E/F) are combined into ``start_at`` / ``stop_at``
``TIMESTAMPTZ`` using ``plants.timezone`` when present, otherwise UTC.
``production_date`` is the Excel date as a DATE (Q1 midnight attribution TBC —
no overnight / +24h invent).

Idempotency
-----------
``external_row_key`` =
``dpr_oee:{plant_id}:{production_date}:{shift_code}:{machine_code}:{part_code}:{start_at_iso}``
(partial UNIQUE on ``production_records.external_row_key``). Re-import updates
the existing raw row, replaces child events, and re-persists metrics.

Does not create APIs, rollups, masters, or Migration 016. Does not resolve
Q1/Q2/Q6/Q11/Q13/Q17. Does not invent downtime/rejection reasons.
"""

from __future__ import annotations

import csv
import io
from collections.abc import Mapping, Sequence
from dataclasses import dataclass, field
from datetime import date, datetime, time
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, BinaryIO
from uuid import UUID
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import delete, select
from sqlalchemy.orm import Session

from app.models.downtime_event import DowntimeEvent
from app.models.downtime_reason import DowntimeReason
from app.models.import_job import ImportJob
from app.models.import_job_row import ImportJobRow
from app.models.machine import Machine
from app.models.operator import Operator
from app.models.part import Part
from app.models.plant import Plant
from app.models.production_record import ProductionRecord
from app.models.rejection_event import RejectionEvent
from app.models.rejection_reason import RejectionReason
from app.models.shift import Shift
from app.services.oee_persistence import persist_production_record_metrics

SHEET_NAME = "DPR_OEE"
HEADER_ROW = 3
SUBHEADER_ROW = 4
DATA_START_ROW = 5

# Built-in DPR_OEE column map (Excel letters → logical field / reason bucket).
COL_DATE = "B"
COL_SHIFT = "C"
COL_MACHINE = "D"
COL_START = "E"
COL_STOP = "F"
COL_OPERATOR = "H"
COL_PART_NAME = "I"
COL_PART_NO = "J"
COL_CAVITY = "K"
COL_CYCLE = "L"
COL_PRODUCED = "N"
COL_PLANNED_DT = "O"
COL_REMARKS = "AV"

# Idle reasons Q–AA (codes 1–11). Labels match Excel row 4 (normalized).
DOWNTIME_COLUMNS: tuple[tuple[str, str, str], ...] = (
    ("Q", "1", "Manpower Shortage"),
    ("R", "2", "Mould Trial"),
    ("S", "3", "Bin Shortage"),
    ("T", "4", "Material Shortage"),
    ("U", "5", "M/c Under BD"),
    ("V", "6", "Nozzle Block"),
    ("W", "7", "Mould Problem"),
    ("X", "8", "Crystal/ Insert Shortage"),
    ("Y", "9", "Power Failure"),
    ("Z", "10", "Process Setting"),
    ("AA", "11", "Others"),
)

# Rejection reasons AH–AQ (codes A–J).
REJECTION_COLUMNS: tuple[tuple[str, str], ...] = (
    ("AH", "A"),
    ("AI", "B"),
    ("AJ", "C"),
    ("AK", "D"),
    ("AL", "E"),
    ("AM", "F"),
    ("AN", "G"),
    ("AO", "H"),
    ("AP", "I"),
    ("AQ", "J"),
)

# Expected primary headers on row 3 (subset — merged group titles for Q/AH OK).
_EXPECTED_ROW3: Mapping[str, str] = {
    "B": "Date",
    "C": "Shift",
    "D": "Machine Name/No.",
    "E": "Start Time",
    "F": "Stop Time",
    "H": "Operator Name",
    "I": "Part Name",
    "J": "Part No.",
    "K": "Cavity",
    "L": "Cycle Time (Sec.)",
    "N": "Prod. Qty. (Pcs.)",
    "O": "Planned Down Time (Tea/Lunch)",
}

_ZERO = Decimal("0")


@dataclass(frozen=True, slots=True)
class ImportJobResult:
    """Outcome of one ``ingest_dpr_oee_workbook`` call (no commit)."""

    import_job_id: UUID
    status: str
    row_count: int
    success_count: int
    error_count: int
    skipped_count: int
    production_record_ids: list[UUID] = field(default_factory=list)
    error_summary: str | None = None


@dataclass
class _ParsedRow:
    excel_row: int
    payload: dict[str, Any]
    production_date: date | None = None
    shift_code: str | None = None
    machine_code: str | None = None
    operator_raw: str | None = None
    part_code: str | None = None
    part_name: str | None = None
    start_time: time | None = None
    stop_time: time | None = None
    cavity_count: Decimal | None = None
    cycle_time_sec: Decimal | None = None
    produced_qty: Decimal | None = None
    planned_downtime_min: Decimal | None = None
    remarks: str | None = None
    downtime_minutes: dict[str, Decimal] = field(default_factory=dict)  # excel_col → min
    rejection_qtys: dict[str, Decimal] = field(default_factory=dict)  # excel_col → qty
    errors: list[dict[str, str]] = field(default_factory=list)


def _col_index(letter: str) -> int:
    """1-based Excel column index from letter (A=1, AA=27)."""
    letter = letter.upper()
    n = 0
    for ch in letter:
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n


def _cell(ws: Worksheet, row: int, letter: str) -> Any:
    return ws.cell(row=row, column=_col_index(letter)).value


def _norm_header(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().split())


def _open_workbook(file_path_or_bytes: str | Path | bytes | BinaryIO) -> Workbook:
    if isinstance(file_path_or_bytes, str | Path):
        return load_workbook(filename=str(file_path_or_bytes), data_only=False)
    if isinstance(file_path_or_bytes, bytes | bytearray):
        return load_workbook(filename=io.BytesIO(file_path_or_bytes), data_only=False)
    # file-like
    return load_workbook(filename=file_path_or_bytes, data_only=False)


@dataclass(frozen=True, slots=True)
class _CsvCellValue:
    """Duck-types openpyxl's ``Cell`` far enough for ``_cell()`` (``.value`` only)."""

    value: Any


class _CsvSheet:
    """Minimal stand-in for an openpyxl ``Worksheet``, backed by parsed CSV rows.

    Implements only the surface the row-parsing/validation code actually uses:
    ``.title``, ``.max_row``, ``.cell(row=, column=).value``. This lets
    ``_cell()``, ``validate_dpr_oee_sheet()``, ``_parse_data_row()``, and the
    row-loop in ``_run_ingestion_rows()`` run unmodified against a CSV export
    of the same DPR_OEE template layout (3 header rows + data from row 5),
    instead of an .xlsx workbook.

    CSV has no sheet-name concept, so ``.title`` is fixed to ``SHEET_NAME`` —
    there is nothing meaningful to validate there for CSV (unlike Excel, where
    a wrong/renamed sheet is a real user error worth catching).
    """

    def __init__(self, rows: list[list[str]]) -> None:
        self._rows = rows
        self.title = SHEET_NAME
        self.max_row = len(rows)

    def cell(self, row: int, column: int) -> _CsvCellValue:
        row_values = self._rows[row - 1] if 0 < row <= len(self._rows) else []
        value: str | None = row_values[column - 1] if 0 < column <= len(row_values) else None
        # Normalise CSV's "" to None so _to_date/_to_time/_to_decimal/_to_str
        # treat a blank cell exactly like a blank Excel cell.
        if value == "":
            value = None
        return _CsvCellValue(value)


def _open_csv_sheet(content: bytes) -> _CsvSheet:
    """Parse CSV bytes (RFC4180 quoting, CRLF or LF, UTF-8 with optional BOM).

    Uses the stdlib ``csv`` module — correctly handles commas inside quoted
    fields (e.g. a free-text Remarks cell) — not a naive ``line.split(",")``.
    """
    text = content.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text))
    rows = [list(raw_row) for raw_row in reader]
    return _CsvSheet(rows)


def _file_uri(file_path_or_bytes: str | Path | bytes | BinaryIO) -> str | None:
    if isinstance(file_path_or_bytes, Path):
        return str(file_path_or_bytes)
    if isinstance(file_path_or_bytes, str):
        return file_path_or_bytes
    return None


def _resolve_tz(plant: Plant) -> ZoneInfo:
    """Plant timezone if valid; else UTC (documented fallback)."""
    name = (plant.timezone or "").strip() or "UTC"
    try:
        return ZoneInfo(name)
    except ZoneInfoNotFoundError:
        return ZoneInfo("UTC")


def _to_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        # ISO date or datetime
        try:
            return date.fromisoformat(text[:10])
        except ValueError:
            return None
    return None


def _to_time(value: Any) -> time | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.time().replace(microsecond=0)
    if isinstance(value, time):
        return value.replace(microsecond=0)
    if isinstance(value, int | float | Decimal):
        # Excel serial fraction of day
        try:
            fraction = float(value)
            if 0 <= fraction < 1:
                total_seconds = int(round(fraction * 86400))
                h, rem = divmod(total_seconds, 3600)
                m, s = divmod(rem, 60)
                return time(h % 24, m, s)
        except (ValueError, OverflowError):
            return None
        return None
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        for fmt in ("%H:%M:%S", "%H:%M"):
            try:
                return datetime.strptime(text, fmt).time()
            except ValueError:
                continue
    return None


def _to_decimal(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, bool):
        return None
    if isinstance(value, int | float):
        return Decimal(str(value))
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        try:
            return Decimal(text)
        except InvalidOperation:
            return None
    return None


def _to_str(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _jsonable(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str | int | float | bool):
        return value
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat()
    return str(value)


def validate_dpr_oee_sheet(ws: Worksheet) -> list[str]:
    """Return workbook-level errors (empty list if sheet/headers look valid)."""
    errors: list[str] = []
    if ws.title != SHEET_NAME:
        errors.append(f"Expected sheet name {SHEET_NAME!r}, got {ws.title!r}")
        return errors

    for letter, expected in _EXPECTED_ROW3.items():
        actual = _norm_header(_cell(ws, HEADER_ROW, letter))
        expected_n = _norm_header(expected)
        # Part Name header in Excel has trailing space — compare loosely.
        if expected_n.lower() not in actual.lower() and actual.lower() not in expected_n.lower():
            errors.append(
                f"Header {letter}{HEADER_ROW}: expected ~{expected!r}, got {actual!r}"
            )

    for letter, _code, _label in DOWNTIME_COLUMNS:
        actual = _norm_header(_cell(ws, SUBHEADER_ROW, letter))
        if not actual:
            errors.append(f"Missing downtime sub-header at {letter}{SUBHEADER_ROW}")

    for letter, _code in REJECTION_COLUMNS:
        actual = _norm_header(_cell(ws, SUBHEADER_ROW, letter))
        if not actual:
            errors.append(f"Missing rejection sub-header at {letter}{SUBHEADER_ROW}")

    return errors


def _is_empty_business_row(parsed: _ParsedRow) -> bool:
    """Skip template rows that lack business inputs (even if formulas exist)."""
    has_identity = any(
        [
            parsed.production_date is not None,
            parsed.shift_code,
            parsed.machine_code,
            parsed.part_code,
            parsed.operator_raw,
            parsed.start_time is not None,
            parsed.stop_time is not None,
            parsed.cavity_count is not None,
            parsed.cycle_time_sec is not None,
            parsed.produced_qty is not None,
            parsed.planned_downtime_min is not None,
            any(v != _ZERO for v in parsed.downtime_minutes.values()),
            any(v != _ZERO for v in parsed.rejection_qtys.values()),
            parsed.remarks,
        ]
    )
    return not has_identity


def _parse_data_row(ws: Worksheet, row_number: int) -> _ParsedRow:
    payload: dict[str, Any] = {"excel_row": row_number}
    for letter in (
        COL_DATE,
        COL_SHIFT,
        COL_MACHINE,
        COL_START,
        COL_STOP,
        COL_OPERATOR,
        COL_PART_NAME,
        COL_PART_NO,
        COL_CAVITY,
        COL_CYCLE,
        COL_PRODUCED,
        COL_PLANNED_DT,
        COL_REMARKS,
    ):
        payload[letter] = _jsonable(_cell(ws, row_number, letter))

    downtime: dict[str, Decimal] = {}
    for letter, _code, _label in DOWNTIME_COLUMNS:
        raw = _cell(ws, row_number, letter)
        payload[letter] = _jsonable(raw)
        # Blank → 0 for SUM semantics; events skipped later when zero.
        minutes = _to_decimal(raw)
        downtime[letter] = _ZERO if minutes is None else minutes

    rejection: dict[str, Decimal] = {}
    for letter, _code in REJECTION_COLUMNS:
        raw = _cell(ws, row_number, letter)
        payload[letter] = _jsonable(raw)
        qty = _to_decimal(raw)
        rejection[letter] = _ZERO if qty is None else qty

    return _ParsedRow(
        excel_row=row_number,
        payload=payload,
        production_date=_to_date(_cell(ws, row_number, COL_DATE)),
        shift_code=_to_str(_cell(ws, row_number, COL_SHIFT)),
        machine_code=_to_str(_cell(ws, row_number, COL_MACHINE)),
        operator_raw=_to_str(_cell(ws, row_number, COL_OPERATOR)),
        part_code=_to_str(_cell(ws, row_number, COL_PART_NO)),
        part_name=_to_str(_cell(ws, row_number, COL_PART_NAME)),
        start_time=_to_time(_cell(ws, row_number, COL_START)),
        stop_time=_to_time(_cell(ws, row_number, COL_STOP)),
        cavity_count=_to_decimal(_cell(ws, row_number, COL_CAVITY)),
        cycle_time_sec=_to_decimal(_cell(ws, row_number, COL_CYCLE)),
        produced_qty=_to_decimal(_cell(ws, row_number, COL_PRODUCED)),
        planned_downtime_min=_to_decimal(_cell(ws, row_number, COL_PLANNED_DT)),
        remarks=_to_str(_cell(ws, row_number, COL_REMARKS)),
        downtime_minutes=downtime,
        rejection_qtys=rejection,
    )


def _add_error(parsed: _ParsedRow, field_name: str, message: str) -> None:
    parsed.errors.append({"field": field_name, "message": message})


def _combine_timestamp(
    production_date: date, t: time, tz: ZoneInfo
) -> datetime:
    """Naive local wall time in plant TZ → aware TIMESTAMPTZ (no +24h invent)."""
    return datetime(
        production_date.year,
        production_date.month,
        production_date.day,
        t.hour,
        t.minute,
        t.second,
        tzinfo=tz,
    )


def _external_row_key(
    *,
    plant_id: UUID,
    production_date: date,
    shift_code: str,
    machine_code: str,
    part_code: str,
    start_at: datetime,
) -> str:
    return (
        f"dpr_oee:{plant_id}:{production_date.isoformat()}:"
        f"{shift_code}:{machine_code}:{part_code}:{start_at.isoformat()}"
    )


def _load_master_maps(
    session: Session, plant_id: UUID
) -> tuple[
    dict[str, Machine],
    dict[str, Shift],
    dict[str, Part],
    list[Operator],
    dict[str, DowntimeReason],
    dict[str, DowntimeReason],
    dict[str, RejectionReason],
    dict[str, RejectionReason],
]:
    machines = {
        m.code: m
        for m in session.scalars(
            select(Machine).where(Machine.plant_id == plant_id)
        ).all()
    }
    shifts = {
        s.code: s
        for s in session.scalars(select(Shift).where(Shift.plant_id == plant_id)).all()
    }
    parts = {p.code: p for p in session.scalars(select(Part)).all()}
    operators = list(session.scalars(select(Operator)).all())

    downtime_rows = list(session.scalars(select(DowntimeReason)).all())
    dt_by_col = {
        r.excel_column: r for r in downtime_rows if r.excel_column
    }
    dt_by_code = {r.code: r for r in downtime_rows}

    rejection_rows = list(session.scalars(select(RejectionReason)).all())
    rj_by_col = {
        r.excel_column: r for r in rejection_rows if r.excel_column
    }
    rj_by_code = {r.code: r for r in rejection_rows}

    return (
        machines,
        shifts,
        parts,
        operators,
        dt_by_col,
        dt_by_code,
        rj_by_col,
        rj_by_code,
    )


def _resolve_operator(
    operators: Sequence[Operator], raw: str | None
) -> Operator | None:
    if not raw:
        return None
    needle = raw.strip().casefold()
    for op in operators:
        if op.name.strip().casefold() == needle:
            return op
    for op in operators:
        if op.employee_code.strip().casefold() == needle:
            return op
    return None


def _validate_and_resolve(
    parsed: _ParsedRow,
    *,
    plant: Plant,
    tz: ZoneInfo,
    machines: Mapping[str, Machine],
    shifts: Mapping[str, Shift],
    parts: Mapping[str, Part],
    operators: Sequence[Operator],
    dt_by_col: Mapping[str, DowntimeReason],
    dt_by_code: Mapping[str, DowntimeReason],
    rj_by_col: Mapping[str, RejectionReason],
    rj_by_code: Mapping[str, RejectionReason],
) -> dict[str, Any] | None:
    """Validate a business row. Returns resolved context or None on errors."""

    if parsed.production_date is None:
        _add_error(parsed, "B", "Date is required")
    if not parsed.shift_code:
        _add_error(parsed, "C", "Shift is required")
    if not parsed.machine_code:
        _add_error(parsed, "D", "Machine Name/No. is required")
    if parsed.start_time is None:
        _add_error(parsed, "E", "Start Time is required")
    if parsed.stop_time is None:
        _add_error(parsed, "F", "Stop Time is required")
    if not parsed.part_code:
        _add_error(parsed, "J", "Part No. is required")
    if parsed.cavity_count is None:
        _add_error(parsed, "K", "Cavity is required")
    elif parsed.cavity_count < _ZERO:
        _add_error(parsed, "K", "Cavity cannot be negative")
    if parsed.cycle_time_sec is None:
        _add_error(parsed, "L", "Cycle Time is required")
    elif parsed.cycle_time_sec < _ZERO:
        _add_error(parsed, "L", "Cycle Time cannot be negative")
    if parsed.produced_qty is None:
        _add_error(parsed, "N", "Prod. Qty. is required")
    elif parsed.produced_qty < _ZERO:
        _add_error(parsed, "N", "Prod. Qty. cannot be negative")

    planned = parsed.planned_downtime_min
    if planned is None:
        planned = _ZERO
    elif planned < _ZERO:
        _add_error(parsed, "O", "Planned Downtime cannot be negative")

    for letter, minutes in parsed.downtime_minutes.items():
        if minutes < _ZERO:
            _add_error(parsed, letter, "Idle minutes cannot be negative")

    for letter, qty in parsed.rejection_qtys.items():
        if qty < _ZERO:
            _add_error(parsed, letter, "Rejection qty cannot be negative")

    machine = machines.get(parsed.machine_code) if parsed.machine_code else None
    if parsed.machine_code and machine is None:
        _add_error(
            parsed,
            "D",
            f"Unknown machine code {parsed.machine_code!r} for plant "
            f"(do not invent masters)",
        )

    shift = shifts.get(parsed.shift_code) if parsed.shift_code else None
    if parsed.shift_code and shift is None:
        _add_error(
            parsed,
            "C",
            f"Unknown shift code {parsed.shift_code!r} for plant "
            f"(do not invent masters)",
        )

    part = parts.get(parsed.part_code) if parsed.part_code else None
    if parsed.part_code and part is None:
        _add_error(
            parsed,
            "J",
            f"Unknown part code {parsed.part_code!r} (do not invent masters)",
        )

    operator: Operator | None = None
    if parsed.operator_raw:
        operator = _resolve_operator(operators, parsed.operator_raw)
        if operator is None:
            _add_error(
                parsed,
                "H",
                f"Unknown operator {parsed.operator_raw!r} "
                f"(match name or employee_code; do not invent masters)",
            )

    resolved_dt: list[tuple[DowntimeReason, Decimal]] = []
    for letter, code, _label in DOWNTIME_COLUMNS:
        minutes = parsed.downtime_minutes.get(letter, _ZERO)
        reason = dt_by_col.get(letter) or dt_by_code.get(code)
        if reason is None:
            # Missing catalog is always an error (even for zero) so fixtures/
            # environments without seeds fail loudly rather than inventing.
            _add_error(
                parsed,
                letter,
                f"Downtime reason for column {letter} / code {code} "
                f"not found in downtime_reasons (do not invent)",
            )
            continue
        if minutes > _ZERO:
            resolved_dt.append((reason, minutes))

    resolved_rj: list[tuple[RejectionReason, Decimal]] = []
    for letter, code in REJECTION_COLUMNS:
        qty = parsed.rejection_qtys.get(letter, _ZERO)
        reason = rj_by_col.get(letter) or rj_by_code.get(code)
        if reason is None:
            _add_error(
                parsed,
                letter,
                f"Rejection reason for column {letter} / code {code} "
                f"not found in rejection_reasons (do not invent)",
            )
            continue
        if qty > _ZERO:
            resolved_rj.append((reason, qty))

    total_rejection = sum((q for _, q in resolved_rj), _ZERO)
    if (
        parsed.produced_qty is not None
        and parsed.produced_qty >= _ZERO
        and total_rejection > parsed.produced_qty
    ):
        _add_error(
            parsed,
            "AH:AQ",
            "Total rejection qty cannot exceed produced qty",
        )

    # Soft check: non-zero idle vs (shift - planned) when start/stop same-day.
    if (
        parsed.errors == []
        and parsed.production_date is not None
        and parsed.start_time is not None
        and parsed.stop_time is not None
    ):
        start_at = _combine_timestamp(parsed.production_date, parsed.start_time, tz)
        stop_at = _combine_timestamp(parsed.production_date, parsed.stop_time, tz)
        if stop_at >= start_at:
            shift_min = Decimal(str((stop_at - start_at).total_seconds())) / Decimal(
                "60"
            )
            available = shift_min - planned
            total_idle = sum((m for _, m in resolved_dt), _ZERO)
            if total_idle > available:
                _add_error(
                    parsed,
                    "Q:AA",
                    "Total idle minutes cannot exceed available time "
                    f"(idle={total_idle}, available={available})",
                )

    if parsed.errors:
        return None

    assert parsed.production_date is not None
    assert parsed.start_time is not None
    assert parsed.stop_time is not None
    assert machine is not None
    assert shift is not None
    assert part is not None
    assert parsed.cavity_count is not None
    assert parsed.cycle_time_sec is not None
    assert parsed.produced_qty is not None
    assert parsed.shift_code is not None
    assert parsed.machine_code is not None
    assert parsed.part_code is not None

    start_at = _combine_timestamp(parsed.production_date, parsed.start_time, tz)
    stop_at = _combine_timestamp(parsed.production_date, parsed.stop_time, tz)
    # Q1: if stop < start on same calendar date, still persist timestamps as-is
    # (no +24h). Calculator will set q1_midnight_unresolved / NULL time metrics.

    return {
        "plant": plant,
        "machine": machine,
        "shift": shift,
        "part": part,
        "operator": operator,
        "production_date": parsed.production_date,
        "start_at": start_at,
        "stop_at": stop_at,
        "cavity_count": parsed.cavity_count,
        "cycle_time_sec": parsed.cycle_time_sec,
        "produced_qty": parsed.produced_qty,
        "planned_downtime_min": planned,
        "remarks": parsed.remarks,
        "downtime": resolved_dt,
        "rejection": resolved_rj,
        "shift_code": parsed.shift_code,
        "machine_code": parsed.machine_code,
        "part_code": parsed.part_code,
        "external_row_key": _external_row_key(
            plant_id=plant.id,
            production_date=parsed.production_date,
            shift_code=parsed.shift_code,
            machine_code=parsed.machine_code,
            part_code=parsed.part_code,
            start_at=start_at,
        ),
    }


def _replace_child_events(
    session: Session,
    production_record: ProductionRecord,
    downtime: Sequence[tuple[DowntimeReason, Decimal]],
    rejection: Sequence[tuple[RejectionReason, Decimal]],
) -> tuple[list[DowntimeEvent], list[RejectionEvent]]:
    """Delete existing events and insert non-zero replacements."""
    session.execute(
        delete(DowntimeEvent).where(
            DowntimeEvent.production_record_id == production_record.id
        )
    )
    session.execute(
        delete(RejectionEvent).where(
            RejectionEvent.production_record_id == production_record.id
        )
    )
    session.flush()

    dt_events: list[DowntimeEvent] = []
    for reason, minutes in downtime:
        ev = DowntimeEvent(
            production_record_id=production_record.id,
            downtime_reason_id=reason.id,
            minutes=minutes,
        )
        session.add(ev)
        dt_events.append(ev)

    rj_events: list[RejectionEvent] = []
    for reason, qty in rejection:
        ev = RejectionEvent(
            production_record_id=production_record.id,
            rejection_reason_id=reason.id,
            qty=qty,
        )
        session.add(ev)
        rj_events.append(ev)

    session.flush()
    return dt_events, rj_events


def _upsert_production_record(
    session: Session,
    *,
    import_job: ImportJob,
    resolved: Mapping[str, Any],
    source_type: str = "excel",
) -> ProductionRecord:
    """Upsert keyed on ``external_row_key`` (business identity — plant/date/
    shift/machine/part/start_at), not on source format. Re-importing the same
    business row via a different format (e.g. CSV after Excel) intentionally
    updates the same record and reassigns source_type/source_import_id —
    same "last wins" duplicate policy already used for repeated Excel imports,
    not a new one.
    """
    key: str = resolved["external_row_key"]
    existing = session.scalar(
        select(ProductionRecord).where(ProductionRecord.external_row_key == key)
    )

    if existing is None:
        record = ProductionRecord(
            plant_id=resolved["plant"].id,
            machine_id=resolved["machine"].id,
            shift_id=resolved["shift"].id,
            operator_id=resolved["operator"].id if resolved["operator"] else None,
            part_id=resolved["part"].id,
            production_date=resolved["production_date"],
            start_at=resolved["start_at"],
            stop_at=resolved["stop_at"],
            cavity_count=resolved["cavity_count"],
            cycle_time_sec=resolved["cycle_time_sec"],
            produced_qty=resolved["produced_qty"],
            planned_downtime_min=resolved["planned_downtime_min"],
            remarks=resolved["remarks"],
            source_import_id=import_job.id,
            source_type=source_type,
            external_row_key=key,
            status="draft",
        )
        session.add(record)
        session.flush()
    else:
        record = existing
        record.plant_id = resolved["plant"].id
        record.machine_id = resolved["machine"].id
        record.shift_id = resolved["shift"].id
        record.operator_id = (
            resolved["operator"].id if resolved["operator"] else None
        )
        record.part_id = resolved["part"].id
        record.production_date = resolved["production_date"]
        record.start_at = resolved["start_at"]
        record.stop_at = resolved["stop_at"]
        record.cavity_count = resolved["cavity_count"]
        record.cycle_time_sec = resolved["cycle_time_sec"]
        record.produced_qty = resolved["produced_qty"]
        record.planned_downtime_min = resolved["planned_downtime_min"]
        record.remarks = resolved["remarks"]
        record.source_import_id = import_job.id
        record.source_type = source_type
        session.flush()

    return record


def _dpr_oee_mapping_config(plant_id: UUID) -> dict[str, Any]:
    return {
        "template": "DPR_OEE",
        "sheet": SHEET_NAME,
        "downtime_columns": [c[0] for c in DOWNTIME_COLUMNS],
        "rejection_columns": [c[0] for c in REJECTION_COLUMNS],
        "plant_id": str(plant_id),
        "timezone_policy": (
            "plants.timezone if resolvable else UTC; "
            "production_date = Excel Date (B); "
            "start_at/stop_at = Date(B)+Time(E/F) in plant TZ; "
            "Q1: no +24h when stop < start"
        ),
    }


def _get_or_init_import_job(
    session: Session,
    *,
    source_type: str,
    resolved_uri: str | None,
    uploaded_by: UUID | None,
    mapping: dict[str, Any],
    import_job: ImportJob | None,
) -> ImportJob:
    """Create a new ``import_jobs`` row, or reset+reuse one (worker/retry path).

    Identical bookkeeping for both Excel and CSV — only ``source_type`` differs.
    """
    if import_job is None:
        import_job = ImportJob(
            source_type=source_type,
            file_uri=resolved_uri,
            uploaded_by=uploaded_by,
            status="validating",
            mapping_config=mapping,
        )
        session.add(import_job)
        session.flush()
        return import_job

    # Reuse existing job (import worker / failed-job retry).
    if resolved_uri is not None:
        import_job.file_uri = resolved_uri
    if uploaded_by is not None:
        import_job.uploaded_by = uploaded_by
    import_job.source_type = source_type
    import_job.status = "validating"
    import_job.mapping_config = {**(import_job.mapping_config or {}), **mapping}
    import_job.row_count = 0
    import_job.success_count = 0
    import_job.error_count = 0
    import_job.error_summary = None
    session.execute(
        delete(ImportJobRow).where(ImportJobRow.import_job_id == import_job.id)
    )
    session.flush()
    return import_job


def _failed_result(import_job: ImportJob, message: str) -> ImportJobResult:
    import_job.status = "failed"
    import_job.error_summary = message
    return ImportJobResult(
        import_job_id=import_job.id,
        status=import_job.status,
        row_count=0,
        success_count=0,
        error_count=0,
        skipped_count=0,
        error_summary=import_job.error_summary,
    )


def _run_ingestion_rows(
    session: Session,
    ws: Any,
    *,
    plant: Plant,
    plant_id: UUID,
    import_job: ImportJob,
    source_type: str,
) -> ImportJobResult:
    """Header validation + row loop + status/counts — shared by Excel and CSV.

    Byte-identical to the pre-CSV single-function body except ``source_type``
    is threaded into ``_upsert_production_record`` instead of hardcoded.
    Operates only through ``_cell()``-compatible ``.cell(row=, column=).value``
    and ``.max_row`` / ``.title``, so it runs unchanged against an openpyxl
    ``Worksheet`` or a ``_CsvSheet``.
    """
    header_errors = validate_dpr_oee_sheet(ws)
    if header_errors:
        result = _failed_result(import_job, "; ".join(header_errors))
        session.flush()
        return result

    (
        machines,
        shifts,
        parts,
        operators,
        dt_by_col,
        dt_by_code,
        rj_by_col,
        rj_by_code,
    ) = _load_master_maps(session, plant_id)
    tz = _resolve_tz(plant)

    success_count = 0
    error_count = 0
    skipped_count = 0
    processed_rows = 0
    production_ids: list[UUID] = []

    max_row = ws.max_row or DATA_START_ROW
    for row_number in range(DATA_START_ROW, max_row + 1):
        parsed = _parse_data_row(ws, row_number)
        if _is_empty_business_row(parsed):
            skipped_count += 1
            continue

        processed_rows += 1
        resolved = _validate_and_resolve(
            parsed,
            plant=plant,
            tz=tz,
            machines=machines,
            shifts=shifts,
            parts=parts,
            operators=operators,
            dt_by_col=dt_by_col,
            dt_by_code=dt_by_code,
            rj_by_col=rj_by_col,
            rj_by_code=rj_by_code,
        )

        if resolved is None:
            error_count += 1
            session.add(
                ImportJobRow(
                    import_job_id=import_job.id,
                    row_number=row_number,
                    raw_row_payload=parsed.payload,
                    validation_errors=list(parsed.errors),
                    external_row_key=None,
                    production_record_id=None,
                )
            )
            continue

        # SAVEPOINT so one row's DB failure does not abort the whole job.
        try:
            with session.begin_nested():
                record = _upsert_production_record(
                    session,
                    import_job=import_job,
                    resolved=resolved,
                    source_type=source_type,
                )
                dt_events, rj_events = _replace_child_events(
                    session,
                    record,
                    resolved["downtime"],
                    resolved["rejection"],
                )
                # Calculator/persistence: timestamps as stored (Q1: no +24h).
                persist_production_record_metrics(
                    session,
                    record,
                    downtime_events=dt_events,
                    rejection_events=rj_events,
                )
                session.flush()
                session.add(
                    ImportJobRow(
                        import_job_id=import_job.id,
                        row_number=row_number,
                        raw_row_payload=parsed.payload,
                        validation_errors=[],
                        external_row_key=resolved["external_row_key"],
                        production_record_id=record.id,
                    )
                )
                session.flush()
                production_ids.append(record.id)
            success_count += 1
        except Exception as exc:  # noqa: BLE001 — row-level failure only
            error_count += 1
            session.add(
                ImportJobRow(
                    import_job_id=import_job.id,
                    row_number=row_number,
                    raw_row_payload=parsed.payload,
                    validation_errors=[
                        {
                            "field": "_persist",
                            "message": f"Persistence failed: {exc}",
                        }
                    ],
                    external_row_key=resolved["external_row_key"],
                    production_record_id=None,
                )
            )

    import_job.row_count = processed_rows
    import_job.success_count = success_count
    import_job.error_count = error_count
    if processed_rows == 0:
        import_job.status = "failed"
        import_job.error_summary = (
            import_job.error_summary
            or "No populated DPR_OEE business rows found (template empty)"
        )
    elif success_count == 0:
        import_job.status = "failed"
        import_job.error_summary = (
            f"{error_count} row(s) failed validation; 0 committed"
        )
    elif error_count > 0:
        import_job.status = "committed"
        import_job.error_summary = (
            f"Partial success: {success_count} ok, {error_count} failed; "
            f"{skipped_count} empty template rows skipped"
        )
    else:
        import_job.status = "committed"
        import_job.error_summary = None

    session.flush()
    return ImportJobResult(
        import_job_id=import_job.id,
        status=import_job.status,
        row_count=import_job.row_count,
        success_count=import_job.success_count,
        error_count=import_job.error_count,
        skipped_count=skipped_count,
        production_record_ids=production_ids,
        error_summary=import_job.error_summary,
    )


def ingest_dpr_oee_workbook(
    session: Session,
    file_path_or_bytes: str | Path | bytes | BinaryIO,
    *,
    plant_id: UUID,
    uploaded_by: UUID | None = None,
    import_job: ImportJob | None = None,
) -> ImportJobResult:
    """Ingest a DPR_OEE Excel workbook into raw + metrics tables (no commit).

    ``plant_id`` is required (Q11: never hard-coded). Masters must already
    exist; missing machine/shift/part/operator/reason → row validation errors.

    When ``import_job`` is provided (worker / retry path), that row is reused
    instead of inserting a new ``import_jobs`` record. Prior ``import_job_rows``
    for the job are cleared so UNIQUE(job, row_number) stays valid.

    Excel-specific: opens the workbook, requires a sheet literally named
    ``DPR_OEE``. Row/validation/persistence logic lives in
    ``_run_ingestion_rows`` (shared with ``ingest_dpr_oee_csv``).
    """
    plant = session.get(Plant, plant_id)
    if plant is None:
        raise ValueError(f"plant_id {plant_id} not found")

    mapping = _dpr_oee_mapping_config(plant_id)
    resolved_uri = _file_uri(file_path_or_bytes)
    import_job = _get_or_init_import_job(
        session,
        source_type="excel",
        resolved_uri=resolved_uri,
        uploaded_by=uploaded_by,
        mapping=mapping,
        import_job=import_job,
    )

    try:
        wb = _open_workbook(file_path_or_bytes)
    except Exception as exc:  # noqa: BLE001 — surface as job failure
        result = _failed_result(import_job, f"Failed to open workbook: {exc}")
        session.flush()
        return result

    if SHEET_NAME not in wb.sheetnames:
        result = _failed_result(
            import_job, f"Sheet {SHEET_NAME!r} not found; sheets={wb.sheetnames!r}"
        )
        session.flush()
        return result

    ws = wb[SHEET_NAME]
    return _run_ingestion_rows(
        session, ws, plant=plant, plant_id=plant_id, import_job=import_job, source_type="excel"
    )


def ingest_dpr_oee_csv(
    session: Session,
    content: bytes,
    *,
    plant_id: UUID,
    uploaded_by: UUID | None = None,
    import_job: ImportJob | None = None,
    file_uri: str | None = None,
) -> ImportJobResult:
    """Ingest a CSV export of the DPR_OEE template into raw + metrics tables.

    CSV Import — Phase 1. Mirrors ``ingest_dpr_oee_workbook`` exactly except
    for how the sheet is opened: this expects a CSV with the *same physical
    layout* as the Excel template (2 title rows, headers on row 3, downtime/
    rejection sub-headers on row 4, data from row 5 — i.e. "Save As CSV" from
    the same workbook), not an arbitrary/generic CSV with one header row.
    A flatter, arbitrarily-columned CSV needs the column-mapping-template
    commit path, which is a later phase (Section 2/§"generalize mapping
    layer" — explicitly out of scope here).

    Same validation, same duplicate policy (external_row_key upsert, no
    source_type in the key — see ``_upsert_production_record``), same
    persistence path as Excel. Only ``ImportJob.source_type`` /
    ``ProductionRecord.source_type`` differ: ``"csv"`` instead of ``"excel"``.
    """
    plant = session.get(Plant, plant_id)
    if plant is None:
        raise ValueError(f"plant_id {plant_id} not found")

    mapping = _dpr_oee_mapping_config(plant_id)
    import_job = _get_or_init_import_job(
        session,
        source_type="csv",
        resolved_uri=file_uri,
        uploaded_by=uploaded_by,
        mapping=mapping,
        import_job=import_job,
    )

    if not content:
        result = _failed_result(import_job, "CSV file is empty")
        session.flush()
        return result

    try:
        ws = _open_csv_sheet(content)
    except Exception as exc:  # noqa: BLE001 — surface as job failure
        result = _failed_result(import_job, f"Failed to parse CSV: {exc}")
        session.flush()
        return result

    if ws.max_row == 0:
        result = _failed_result(import_job, "CSV file is empty")
        session.flush()
        return result
    if ws.max_row < DATA_START_ROW:
        result = _failed_result(
            import_job,
            "CSV has no data rows — expected the DPR_OEE template layout "
            f"(headers on row {HEADER_ROW}, data from row {DATA_START_ROW})",
        )
        session.flush()
        return result

    return _run_ingestion_rows(
        session, ws, plant=plant, plant_id=plant_id, import_job=import_job, source_type="csv"
    )


# Re-export helpers useful for unit tests.
__all__ = [
    "DOWNTIME_COLUMNS",
    "ImportJobResult",
    "REJECTION_COLUMNS",
    "SHEET_NAME",
    "ingest_dpr_oee_csv",
    "ingest_dpr_oee_workbook",
    "validate_dpr_oee_sheet",
]
