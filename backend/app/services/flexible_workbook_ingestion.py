"""Flexible manufacturing workbook ingestion (any sheet format, not just DPR_OEE).

This service complements dpr_oee_ingestion.py and handles Excel/CSV files
with arbitrary manufacturing data structures. It auto-detects column headers,
maps them to canonical fields, and persists to the database without requiring
the rigid DPR_OEE template.

Unlike dpr_oee_ingestion.py:
- No fixed sheet name requirement
- No fixed column letter mappings
- Flexible header detection with aliases
- Works with any reasonable manufacturing data structure
- Forward-compatible with flexible imports from frontend
"""

from __future__ import annotations

import csv
import io
import re
from collections.abc import Mapping
from dataclasses import dataclass, field
from datetime import date, datetime, time
from decimal import Decimal
from pathlib import Path
from typing import Any, BinaryIO
from uuid import UUID

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import delete, select
from sqlalchemy.orm import Session

from app.models.downtime_event import DowntimeEvent
from app.models.downtime_reason import DowntimeReason
from app.models.import_job import ImportJob
from app.models.import_job_row import ImportJobRow
from app.models.machine import Machine
from app.models.operator import Operator
from app.models.part import Part
from app.models.plant import Plant
from app.models.production_record import ProductionRecord
from app.models.rejection_event import RejectionEvent
from app.models.rejection_reason import RejectionReason
from app.models.shift import Shift
from app.services.oee_persistence import persist_production_record_metrics

# Column aliases: normalized header → canonical field name
_CANONICAL_ALIASES: Mapping[str, list[str]] = {
    "date": ["date", "production date", "production_date", "prod date", "shift date", "date of production"],
    "line": ["line", "production line", "prod line", "line name", "line_name"],
    "shift": ["shift", "shift name", "shift_name", "shift code"],
    "machine": ["machine", "machine name", "machine_name", "machine no", "machine_no", "m/c", "mc"],
    "part": ["part", "part name", "part_name", "part no", "part_no", "product", "product name"],
    "production": [
        "production",
        "actual production",
        "actual production qty",
        "prod qty",
        "produced qty",
        "total production",
        "total prod nos",
        "produced pcs",
    ],
    "target": ["target", "production target", "prod target", "target qty", "target nos", "planned qty"],
    "downtime": [
        "downtime",
        "idle time",
        "downtime minutes",
        "downtime mints",
        "total idle time",
        "breakdown minutes",
        "breakdown time",
        "idle minutes",
    ],
    "rejection": [
        "rejection",
        "rejected qty",
        "rejection qty",
        "total rejection",
        "quality loss",
        "defective qty",
        "total rejection (pcs qty.)",
    ],
    "availability": ["availability", "availability ratio", "operating time", "run time", "availability %"],
    "performance": [
        "performance",
        "performance ratio",
        "operator efficiency",
        "machine efficiency",
        "performance %",
    ],
    "quality": ["quality", "quality rate", "quality ratio", "quality %"],
    "oee": ["oee", "oee %", "oee percentage", "overall equipment effectiveness"],
}

_ZERO = Decimal("0")


@dataclass(frozen=True, slots=True)
class FlexibleImportResult:
    """Outcome of flexible workbook ingestion."""

    import_job_id: UUID
    status: str
    sheet_name: str
    row_count: int
    success_count: int
    error_count: int
    skipped_count: int
    production_record_ids: list[UUID] = field(default_factory=list)
    error_summary: str | None = None


def _normalize_header(label: str) -> str:
    """Normalize header to lowercase alphanumeric."""
    cleaned = re.sub(r"[^a-z0-9]+", "", (label or "").strip().lower())
    return cleaned or "column"


def _get_canonical_field(header: str) -> str | None:
    """Map header to canonical field name using aliases."""
    normalized = _normalize_header(header)
    for canonical, aliases in _CANONICAL_ALIASES.items():
        for alias in aliases:
            if normalized == _normalize_header(alias):
                return canonical
    return None


def _coerce_value(value: Any) -> Any:
    """Coerce raw cell value to string or None."""
    if value is None or value == "":
        return None
    if isinstance(value, str):
        stripped = value.strip()
        return stripped if stripped else None
    return str(value).strip() if str(value).strip() else None


def _coerce_date(value: Any) -> str | None:
    """Parse various date formats to ISO string."""
    if value is None or value == "":
        return None

    # If already a datetime
    if isinstance(value, datetime):
        return value.date().isoformat()

    if isinstance(value, date):
        return value.isoformat()

    # String parsing
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return None

        # Try common formats
        for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d", "%d/%m/%y", "%m/%d/%y"):
            try:
                return datetime.strptime(stripped, fmt).date().isoformat()
            except ValueError:
                continue

        # Try ISO format
        try:
            return datetime.fromisoformat(stripped).date().isoformat()
        except ValueError:
            pass

    return None


def _coerce_number(value: Any) -> float | None:
    """Parse number, handling strings with commas/percent."""
    if value is None or value == "":
        return None

    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)

    if isinstance(value, str):
        cleaned = value.strip().replace(",", "").replace("%", "")
        if not cleaned:
            return None
        try:
            return float(cleaned)
        except ValueError:
            return None

    return None


def _select_best_sheet(workbook) -> str:
    """Auto-select the most likely manufacturing data sheet."""
    import openpyxl.utils

    sheet_scores = []
    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        # Sample first 30 rows looking for manufacturing headers
        headers_found = 0
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=30, values_only=True), 1):
            if not row or not any(cell for cell in row):
                continue
            for cell in row:
                if cell:
                    normalized = _normalize_header(str(cell))
                    # Check if any canonical field matches
                    for canonical, aliases in _CANONICAL_ALIASES.items():
                        for alias in aliases:
                            if normalized == _normalize_header(alias):
                                headers_found += 1
                                break
            if headers_found >= 2:
                break

        # Count data rows
        data_rows = 0
        for row in ws.iter_rows(min_row=30, max_row=1000, values_only=True):
            if row and any(cell for cell in row):
                data_rows += 1

        score = headers_found * 100 + min(data_rows, 1000)
        sheet_scores.append((sheet_name, score))

    if not sheet_scores:
        raise ValueError("No sheets found with manufacturing data.")

    best_sheet = max(sheet_scores, key=lambda x: x[1])[0]
    if sheet_scores[0][1] == 0:
        raise ValueError("No analyzable manufacturing data found in workbook.")

    return best_sheet


def _find_header_row(rows: list[list[Any]]) -> int:
    """Find the row index that most likely contains headers."""
    best_idx = -1
    best_score = 0

    for idx, row in enumerate(rows[:40]):  # Check first 40 rows
        if not row:
            continue
        score = 0
        for cell in row:
            if cell:
                normalized = _normalize_header(str(cell))
                for canonical, aliases in _CANONICAL_ALIASES.items():
                    for alias in aliases:
                        if normalized == _normalize_header(alias):
                            score += 1
                            break

        if score > best_score:
            best_score = score
            best_idx = idx

    if best_idx == -1 or best_score < 1:
        raise ValueError("No identifiable headers found in sheet.")

    return best_idx


def ingest_flexible_workbook(
    db: Session,
    file_path: Path | str,
    plant_id: UUID,
    import_job_id: UUID | None = None,
) -> FlexibleImportResult:
    """Ingest Excel/CSV with flexible (non-DPR_OEE) format.

    Auto-detects sheet, headers, column mapping. Normalizes to ProductionRecord.
    """
    file_path = Path(file_path) if isinstance(file_path, str) else file_path

    if not file_path.exists():
        raise FileNotFoundError(f"File not found: {file_path}")

    # Load workbook
    try:
        workbook = load_workbook(file_path, data_only=True)
    except Exception as exc:
        raise ValueError(f"Unable to read Excel file: {exc}") from exc

    # Auto-select best sheet
    sheet_name = _select_best_sheet(workbook)
    ws = workbook[sheet_name]

    # Read all rows
    all_rows = list(ws.iter_rows(values_only=True))

    # Find header row
    header_row_idx = _find_header_row(all_rows)
    header_row = all_rows[header_row_idx]
    data_start_idx = header_row_idx + 1

    # Build header mapping
    headers: list[str | None] = [_normalize_header(str(cell)) if cell else None for cell in header_row]

    # Map headers to canonical fields
    column_map: dict[int, str] = {}  # col_idx → canonical_field
    for col_idx, header in enumerate(headers):
        if header and header != "column":
            canonical = _get_canonical_field(header)
            if canonical:
                column_map[col_idx] = canonical

    if not column_map:
        raise ValueError("No recognized manufacturing columns found.")

    # Parse data rows
    parsed_rows = []
    for row_data in all_rows[data_start_idx:]:
        if not row_data or not any(cell for cell in row_data):
            continue

        record: dict[str, Any] = {}
        for col_idx, canonical_field in column_map.items():
            value = row_data[col_idx] if col_idx < len(row_data) else None

            if canonical_field == "date":
                record[canonical_field] = _coerce_date(value)
            elif canonical_field in {"production", "target", "downtime", "rejection"}:
                record[canonical_field] = _coerce_number(value)
            else:
                record[canonical_field] = _coerce_value(value)

        if any(record.values()):  # At least one non-null field
            parsed_rows.append(record)

    if not parsed_rows:
        raise ValueError("No data rows found.")

    # Create or get import job
    if import_job_id:
        job = db.scalar(select(ImportJob).where(ImportJob.id == import_job_id))
        if not job:
            raise ValueError(f"ImportJob {import_job_id} not found.")
        # Clear previous rows
        db.execute(delete(ImportJobRow).where(ImportJobRow.import_job_id == import_job_id))
        db.flush()
    else:
        job = ImportJob(plant_id=plant_id, status="processing", source="flexible_workbook", source_file=str(file_path))
        db.add(job)
        db.flush()

    # Process records
    plant = db.scalar(select(Plant).where(Plant.id == plant_id))
    if not plant:
        raise ValueError(f"Plant {plant_id} not found.")

    production_record_ids: list[UUID] = []
    success_count = 0
    error_count = 0
    skipped_count = 0

    for row_idx, row_data in enumerate(parsed_rows, start=1):
        try:
            # Extract fields
            prod_date = row_data.get("date")
            if not prod_date:
                skipped_count += 1
                continue

            machine_name = row_data.get("machine") or "Unknown"
            line_name = row_data.get("line") or None
            shift_code = row_data.get("shift") or "A"
            part_name = row_data.get("part") or None
            production_qty = row_data.get("production") or 0
            target_qty = row_data.get("target") or None
            downtime_minutes = row_data.get("downtime") or 0
            rejection_qty = row_data.get("rejection") or 0

            # Get or create masters
            machine = db.scalar(select(Machine).where(Machine.plant_id == plant_id, Machine.name == machine_name))
            if not machine:
                machine = Machine(plant_id=plant_id, name=machine_name)
                db.add(machine)
                db.flush()

            shift = db.scalar(select(Shift).where(Shift.plant_id == plant_id, Shift.code == shift_code))
            if not shift:
                shift = Shift(plant_id=plant_id, code=shift_code, name=shift_code)
                db.add(shift)
                db.flush()

            part = None
            if part_name:
                part = db.scalar(select(Part).where(Part.plant_id == plant_id, Part.name == part_name))
                if not part:
                    part = Part(plant_id=plant_id, name=part_name)
                    db.add(part)
                    db.flush()

            # Create production record
            external_row_key = f"flexible:{plant_id}:{prod_date}:{shift_code}:{machine_name}:{part_name or 'none'}:{row_idx}"

            prod_rec = ProductionRecord(
                import_job_id=job.id,
                plant_id=plant_id,
                production_date=datetime.strptime(prod_date, "%Y-%m-%d").date() if isinstance(prod_date, str) else prod_date,
                shift_id=shift.id,
                machine_id=machine.id,
                part_id=part.id if part else None,
                start_at=None,
                stop_at=None,
                external_row_key=external_row_key,
                source_import_id=str(job.id),
                raw_data={
                    "production": float(production_qty) if production_qty else 0,
                    "target": float(target_qty) if target_qty else None,
                    "downtime_minutes": float(downtime_minutes) if downtime_minutes else 0,
                    "rejection": float(rejection_qty) if rejection_qty else 0,
                },
            )
            db.add(prod_rec)
            db.flush()

            production_record_ids.append(prod_rec.id)

            # Record result
            job_row = ImportJobRow(
                import_job_id=job.id,
                excel_row=header_row_idx + row_idx,
                payload=row_data,
                validation_errors=None,
                production_record_id=prod_rec.id,
            )
            db.add(job_row)
            success_count += 1

        except Exception as exc:
            error_count += 1
            job_row = ImportJobRow(
                import_job_id=job.id,
                excel_row=header_row_idx + row_idx,
                payload=row_data,
                validation_errors=str(exc),
            )
            db.add(job_row)

    # Update job status
    job.status = "completed"
    job.result = f"{success_count} success, {error_count} errors, {skipped_count} skipped"

    return FlexibleImportResult(
        import_job_id=job.id,
        status="completed",
        sheet_name=sheet_name,
        row_count=len(parsed_rows),
        success_count=success_count,
        error_count=error_count,
        skipped_count=skipped_count,
        production_record_ids=production_record_ids,
        error_summary=job.result if error_count > 0 else None,
    )


def ingest_flexible_csv(
    db: Session,
    csv_content: str | bytes,
    plant_id: UUID,
    import_job_id: UUID | None = None,
) -> FlexibleImportResult:
    """Ingest CSV with flexible format."""
    if isinstance(csv_content, bytes):
        csv_content = csv_content.decode("utf-8")

    reader = csv.reader(io.StringIO(csv_content))
    all_rows = list(reader)

    if not all_rows:
        raise ValueError("CSV is empty.")

    # Find header row
    header_row_idx = _find_header_row(all_rows)
    header_row = all_rows[header_row_idx]
    data_start_idx = header_row_idx + 1

    # Build header mapping
    headers: list[str | None] = [_normalize_header(cell) if cell else None for cell in header_row]
    column_map: dict[int, str] = {}
    for col_idx, header in enumerate(headers):
        if header and header != "column":
            canonical = _get_canonical_field(header)
            if canonical:
                column_map[col_idx] = canonical

    if not column_map:
        raise ValueError("No recognized manufacturing columns found.")

    # Parse data rows
    parsed_rows = []
    for row_data in all_rows[data_start_idx:]:
        if not row_data or not any(cell.strip() for cell in row_data if isinstance(cell, str)):
            continue

        record: dict[str, Any] = {}
        for col_idx, canonical_field in column_map.items():
            value = row_data[col_idx] if col_idx < len(row_data) else None

            if canonical_field == "date":
                record[canonical_field] = _coerce_date(value)
            elif canonical_field in {"production", "target", "downtime", "rejection"}:
                record[canonical_field] = _coerce_number(value)
            else:
                record[canonical_field] = _coerce_value(value)

        if any(record.values()):
            parsed_rows.append(record)

    if not parsed_rows:
        raise ValueError("No data rows found in CSV.")

    # Create or get import job
    if import_job_id:
        job = db.scalar(select(ImportJob).where(ImportJob.id == import_job_id))
        if not job:
            raise ValueError(f"ImportJob {import_job_id} not found.")
        db.execute(delete(ImportJobRow).where(ImportJobRow.import_job_id == import_job_id))
        db.flush()
    else:
        job = ImportJob(plant_id=plant_id, status="processing", source="flexible_csv", source_file="uploaded.csv")
        db.add(job)
        db.flush()

    # Process records (same as Excel)
    plant = db.scalar(select(Plant).where(Plant.id == plant_id))
    if not plant:
        raise ValueError(f"Plant {plant_id} not found.")

    production_record_ids: list[UUID] = []
    success_count = 0
    error_count = 0
    skipped_count = 0

    for row_idx, row_data in enumerate(parsed_rows, start=1):
        try:
            prod_date = row_data.get("date")
            if not prod_date:
                skipped_count += 1
                continue

            machine_name = row_data.get("machine") or "Unknown"
            shift_code = row_data.get("shift") or "A"
            part_name = row_data.get("part") or None

            machine = db.scalar(select(Machine).where(Machine.plant_id == plant_id, Machine.name == machine_name))
            if not machine:
                machine = Machine(plant_id=plant_id, name=machine_name)
                db.add(machine)
                db.flush()

            shift = db.scalar(select(Shift).where(Shift.plant_id == plant_id, Shift.code == shift_code))
            if not shift:
                shift = Shift(plant_id=plant_id, code=shift_code, name=shift_code)
                db.add(shift)
                db.flush()

            part = None
            if part_name:
                part = db.scalar(select(Part).where(Part.plant_id == plant_id, Part.name == part_name))
                if not part:
                    part = Part(plant_id=plant_id, name=part_name)
                    db.add(part)
                    db.flush()

            external_row_key = f"flexible_csv:{plant_id}:{prod_date}:{shift_code}:{machine_name}:{part_name or 'none'}:{row_idx}"

            prod_rec = ProductionRecord(
                import_job_id=job.id,
                plant_id=plant_id,
                production_date=datetime.strptime(prod_date, "%Y-%m-%d").date() if isinstance(prod_date, str) else prod_date,
                shift_id=shift.id,
                machine_id=machine.id,
                part_id=part.id if part else None,
                start_at=None,
                stop_at=None,
                external_row_key=external_row_key,
                source_import_id=str(job.id),
                raw_data={
                    "production": float(row_data.get("production") or 0),
                    "target": float(row_data.get("target")) if row_data.get("target") else None,
                    "downtime_minutes": float(row_data.get("downtime") or 0),
                    "rejection": float(row_data.get("rejection") or 0),
                },
            )
            db.add(prod_rec)
            db.flush()

            production_record_ids.append(prod_rec.id)

            job_row = ImportJobRow(
                import_job_id=job.id,
                excel_row=header_row_idx + row_idx,
                payload=row_data,
                validation_errors=None,
                production_record_id=prod_rec.id,
            )
            db.add(job_row)
            success_count += 1

        except Exception as exc:
            error_count += 1
            job_row = ImportJobRow(
                import_job_id=job.id,
                excel_row=header_row_idx + row_idx,
                payload=row_data,
                validation_errors=str(exc),
            )
            db.add(job_row)

    job.status = "completed"
    job.result = f"{success_count} success, {error_count} errors, {skipped_count} skipped"

    return FlexibleImportResult(
        import_job_id=job.id,
        status="completed",
        sheet_name="CSV",
        row_count=len(parsed_rows),
        success_count=success_count,
        error_count=error_count,
        skipped_count=skipped_count,
        production_record_ids=production_record_ids,
        error_summary=job.result if error_count > 0 else None,
    )
